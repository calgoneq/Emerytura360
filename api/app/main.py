from fastapi import FastAPI, Body, Query, Response, HTTPException
from pydantic import BaseModel, Field
from typing import Optional, Dict, List
import datetime as dt
from pathlib import Path
from fastapi.responses import StreamingResponse, RedirectResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic.config import ConfigDict

# stdlib / io
import os
import io
import csv
import copy

# ReportLab
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib import colors

# XLSX
import openpyxl

# local calc engine
from .calculations.engine import (
    efekt_absencji_factor, waloryzuj_rocznie, waloryzuj_kwartalnie_po_31_stycznia,
    annuitetyzuj, urealnij
)
from .calculations.waloryzacja import A as ASSUMPTIONS

# --- optional .env ---
try:
    from dotenv import load_dotenv
    load_dotenv()
except Exception:
    pass

# --- Paths / storage ---
BASE = Path(__file__).resolve().parents[1]
STORAGE = BASE / "storage"; STORAGE.mkdir(exist_ok=True)
FONTS_DIR = BASE / "fonts"; FONTS_DIR.mkdir(exist_ok=True)
DATA_DIR = BASE / "data"; DATA_DIR.mkdir(exist_ok=True)
LOG_CSV = STORAGE / "usage.csv"

# --- Colors (ZUS palette) ---
ZUS_ORANGE = colors.Color(255/255, 179/255, 79/255)
ZUS_GREEN  = colors.Color(0/255, 153/255, 63/255)
ZUS_GRAY   = colors.Color(190/255, 195/255, 206/255)
ZUS_BLUE   = colors.Color(63/255, 132/255, 210/255)
ZUS_NAVY   = colors.Color(0/255, 65/255, 110/255)
ZUS_RED    = colors.Color(240/255, 94/255, 94/255)
ZUS_BLACK  = colors.black
ZUS_LIGHT_BG = colors.Color(246/255, 248/255, 251/255)  # delikatne tło sekcji

# Dodatkowe akcenty do „annual report”
ACCENT_YELLOW = colors.Color(255/255, 203/255, 0/255)
ACCENT_TEAL   = colors.Color(0/255, 185/255, 185/255)

# --- Layout constants for PDF ---
MARGIN_X = 32        # lewy/prawy margines strony
MARGIN_TOP = 78      # górny margines (dla startu treści)
SECTION_GAP = 22     # odstęp pionowy między sekcjami
LINE_GAP = 14        # odstęp między linijkami tekstu

# --- Typography (CSS-like font stack dla ReportLab) ---
FONT_MAIN = "Helvetica"          # ostateczny fallback (Type1)
FONT_BOLD = "Helvetica-Bold"

def _register_polish_fonts():
    """
    Ustawia główną rodzinę czcionek wg kolejności jak w CSS:
    ui-sans-serif, system-ui, -apple-system, 'Segoe UI', Roboto, Helvetica, Arial.
    Szuka plików TTF/OTF w katalogu ./fonts oraz w typowych ścieżkach systemowych.
    Pierwsza znaleziona para (Regular + Bold) staje się FONT_MAIN / FONT_BOLD.
    Zawsze zostawia Helvetica jako ostateczny fallback.
    """
    import sys
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    # 1) Kolejność preferencji (nazwa logiczna, lista wzorców plików Regular, Bold)
    PREFERRED = [
        ("SFPro",              # -apple-system / system-ui na macOS
         ["SFProText-Regular.otf", "SFUIText-Regular.otf", "SFNS.ttf"],
         ["SFProText-Bold.otf", "SFUIText-Bold.otf", "SFNS-Bold.ttf"]),
        ("SegoeUI",            # Windows (Segoe UI)
         ["segoeui.ttf", "Segoe UI.ttf"],
         ["segoeuib.ttf", "Segoe UI Bold.ttf"]),
        ("Roboto",             # Android / web
         ["Roboto-Regular.ttf"],
         ["Roboto-Bold.ttf"]),
        ("Arial",
         ["Arial.ttf", "arial.ttf"],
         ["Arial Bold.ttf", "arialbd.ttf"]),
        # Bardzo dobry fallback z pełnymi znakami PL (dołącz jeśli chcesz)
        ("DejaVuSans",
         ["DejaVuSans.ttf"],
         ["DejaVuSans-Bold.ttf"]),
    ]

    # 2) Gdzie szukać plików
    search_dirs = [
        str(FONTS_DIR),                         # ./fonts w projekcie
        "/System/Library/Fonts", "/Library/Fonts",                       # macOS
        "C:/Windows/Fonts",                                                         # Windows
        "/usr/share/fonts", "/usr/local/share/fonts", "~/.local/share/fonts", "~/.fonts"  # Linux
    ]
    search_dirs = [os.path.expanduser(d) for d in search_dirs if os.path.isdir(os.path.expanduser(d))]

    def _find_first(paths):
        for base in search_dirs:
            for name in paths:
                p = os.path.join(base, name)
                if os.path.isfile(p):
                    return p
        return None

    def _try_register(alias, regular_candidates, bold_candidates):
        reg = _find_first(regular_candidates)
        bold = _find_first(bold_candidates)
        if not reg or not bold:
            return False
        try:
            pdfmetrics.registerFont(TTFont(alias, reg))
            pdfmetrics.registerFont(TTFont(alias+"-Bold", bold))
            return True
        except Exception:
            return False

    # 3) Przejdź po stacku i wybierz pierwszą działającą parę
    global FONT_MAIN, FONT_BOLD
    for alias, regs, bolds in PREFERRED:
        if _try_register(alias, regs, bolds):
            FONT_MAIN = alias
            FONT_BOLD = alias + "-Bold"
            break
    # jeśli nic nie znaleziono, zostaje Helvetica/Helvetica-Bold
    # (wtedy diakrytyki PL mogą być gorsze – najlepiej wrzuć Roboto/DejaVu do ./fonts)

_register_polish_fonts()

# --- App ---
app = FastAPI(title="Emerytura360 API", version="0.4.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.get("/", include_in_schema=False)
def root():
    return RedirectResponse(url="/docs")

@app.get("/health")
def health():
    return {
        "ok": True,
        "version": "0.4.0",
        "avg_loaded": bool(AVG_TABLE),
        "params_loaded": bool(PARAMS),
        "demo": DEMO
    }

# --- MODELS ---
class Balance(BaseModel):
    konto: Optional[float] = 0.0
    subkonto: Optional[float] = 0.0

class SimInput(BaseModel):
    age: int = Field(..., ge=16, le=80)
    sex: str = Field(..., pattern="^[KkMm]$")
    gross_salary: float = Field(..., gt=0)
    start_year: int = Field(
        ...,
        description="Rok rozpoczęcia pracy (liczony od stycznia danego roku)."
    )
    retire_year: Optional[int] = Field(
        None,
        description="Planowany rok zakończenia aktywności zawodowej (styczeń). Domyślnie: osiągnięcie wieku 60 (K) / 65 (M)."
    )
    include_sick_leave: bool = Field(
        True,
        description="Uwzględnia średnią absencję chorobową (dni/rok) różną dla K/M."
    )
    quarter_award: int = Field(3, ge=1, le=4)
    zus_balance: Optional[Balance] = None
    custom_wage_timeline: Optional[Dict[int, float]] = None
    custom_sick_days: Optional[Dict[int, float]] = None  # ⬅️ DODAJ TO
    expected_pension: Optional[float] = None
    postal_code: Optional[str] = None

    model_config = ConfigDict(json_schema_extra={
        "example": {
            "age": 28, "sex": "K", "gross_salary": 8500,
            "start_year": 2020, "retire_year": 2065,
            "include_sick_leave": True, "quarter_award": 3,
            "zus_balance": {"konto": 0, "subkonto": 0},
            "custom_wage_timeline": None,
            "expected_pension": 5000, "postal_code": "30-001"
        }
    })

# --- Helpers: assumptions & averages table ---
AVG_TABLE: Dict[int, float] = {}

# --- Mentor params (CPI, real wage, avg wage, waloryzacje) ---
PARAMS: Dict[int, dict] = {}

def _norm(s) -> str:
    return str(s or "").replace("\xa0", " ").strip().lower()

def _find_col(headers: Dict[str, int], *fragments: str) -> Optional[int]:
    for name, idx in headers.items():
        n = _norm(name)
        if all(_norm(frag) in n for frag in fragments):
            return idx
    return None

def _to_float(v) -> Optional[float]:
    try:
        s = str(v).replace(" ", "").replace("\xa0", "").replace(",", ".")
        return float(s)
    except Exception:
        return None

def load_params_table():
    """
    Wczytuje arkusz mentorów: data/parametry_mentor.xlsx (aktywny arkusz).
    Oczekiwane kolumny (wystarczy fragment nazwy):
    - 'rok'
    - 'wskaźnik cen towarów i usług' (CPI, np. 1.0360)
    - 'realnego wzrostu przeciętn' (real wage index, np. 1.0202)
    - 'przeciętne miesięczne wynagrodzenie'
    - 'waloryzacji składek ... na koncie'
    - 'waloryzacji składek ... na subkoncie'
    """
    path = DATA_DIR / "parametry_mentor.xlsx"
    if not path.exists():
        return
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        headers = { (ws.cell(row=1, column=i).value or ""): i for i in range(1, ws.max_column+1) }

        c_year = _find_col(headers, "rok")
        c_cpi  = _find_col(headers, "wskaźnik cen towarów i usług")
        c_rw   = _find_col(headers, "realnego wzrostu przeciętn")  # real wage
        c_avg  = _find_col(headers, "przeciętne miesięczne wynagrodzenie")
        c_wk   = _find_col(headers, "waloryzacji", "na koncie")
        c_ws   = _find_col(headers, "waloryzacji", "na subkoncie")

        for r in range(2, ws.max_row+1):
            y_raw = ws.cell(row=r, column=c_year).value if c_year else None
            if not y_raw:
                continue
            try:
                y = int(str(y_raw).split(".")[0])
            except Exception:
                continue

            cpi_idx   = _to_float(ws.cell(r, c_cpi).value) if c_cpi else None
            real_wage = _to_float(ws.cell(r, c_rw).value) if c_rw else None
            avg_wage  = _to_float(ws.cell(r, c_avg).value) if c_avg else None
            wal_konto = _to_float(ws.cell(r, c_wk).value) if c_wk else None
            wal_sub   = _to_float(ws.cell(r, c_ws).value) if c_ws else None

            PARAMS[y] = {
                "cpi_index": cpi_idx,           # np. 1.036 -> CPI=3.6%
                "real_wage_index": real_wage,   # np. 1.0202
                "avg_wage": avg_wage,           # PLN/m-c
                "wal_konto": wal_konto,         # np. 114.41% -> 114.41
                "wal_sub": wal_sub,             # np. 109.83% -> 109.83
            }
    except Exception:
        # nie psuj uruchomienia, po prostu zostaw pusty PARAMS
        pass

load_params_table()

def load_avg_benefit_table():
    """
    Ładuje średnie emerytury z pliku XLSX (api/data/avg_benefit.xlsx):
    kolumny: 'rok', 'kwota' (w PLN/m-c).
    """
    path = DATA_DIR / "avg_benefit.xlsx"
    if not path.exists():
        return
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        headers = { (ws.cell(row=1, column=i).value or "").strip().lower(): i
                    for i in range(1, ws.max_column+1) }
        c_year = headers.get("rok")
        c_val  = headers.get("kwota")
        if not c_year or not c_val:
            c_year, c_val = 1, 2
        for r in range(2, ws.max_row+1):
            y = ws.cell(row=r, column=c_year).value
            v = ws.cell(row=r, column=c_val).value
            try:
                y = int(str(y).split(".")[0])
                v = float(str(v).replace(" ", "").replace("\xa0","").replace(",", "."))
                if 1990 <= y <= 2100 and v > 0:
                    AVG_TABLE[y] = v
            except Exception:
                continue
    except Exception:
        pass

load_avg_benefit_table()

# DEMO mode (np. DEMO=1 w .env)
DEMO = os.getenv("DEMO", "0") == "1"

def _seed_avg_if_missing():
    """Prosty seed średnich dla dema/hackathonu, gdy brak pliku avg_benefit.xlsx."""
    if not AVG_TABLE:
        base = 3800.0
        for y in range(dt.date.today().year, dt.date.today().year + 10):
            # prosty trend 3% rocznie
            AVG_TABLE[y] = base * (1.03 ** (y - dt.date.today().year))

# W trybie demo zawsze dosiej średnią, jeśli brak
if DEMO:
    _seed_avg_if_missing()

def _fallback_growth() -> float:
    try:
        return float(os.getenv("AVERAGES_FALLBACK_GROWTH", "0.03"))
    except Exception:
        return 0.03
    
def wage_growth_rate() -> float:
    """
    Średni roczny wzrost wynagrodzeń do projekcji płacy do roku przejścia.
    Hackathon: brane z ENV WAGE_GROWTH (domyślnie 0.03).
    """
    try:
        return float(os.getenv("WAGE_GROWTH", "0.03"))
    except Exception:
        return 0.03

def statutory_retire_age(sex: str) -> int:
    """Ustawowy wiek emerytalny: 60 lat K, 65 lat M."""
    return 60 if str(sex).upper() == "K" else 65

def _avg_growth_rate_from_tail(data: Dict[int, float], k: int = 5) -> float:
    if not data or len(data) < 2:
        return _fallback_growth()
    items = sorted(data.items())
    rates: List[float] = []
    for i in range(1, min(len(items), k + 1)):
        if len(items) - i - 1 < 0:
            break
        y2, v2 = items[-i]
        y1, v1 = items[-i - 1]
        if v1 and v1 > 0 and y2 > y1:
            r = (v2 / v1) ** (1.0 / (y2 - y1)) - 1.0
            rates.append(r)
    return sum(rates) / len(rates) if rates else _fallback_growth()

def expected_life_months(sex: str, retire_year: int) -> int:
    # TODO: tablice GUS / e_x
    return 240

def absencja_days(sex: str) -> Optional[float]:
    key = "K" if sex.upper() == "K" else "M"
    return ASSUMPTIONS.get("absencja_chorobowa", {}).get(key, {}).get("dni_rocznie")

def ensure_log_header():
    if not LOG_CSV.exists():
        with LOG_CSV.open("w", newline="", encoding="utf-8") as f:
            csv.writer(f).writerow([
                "date","time","expected_pension","age","sex","salary",
                "included_sick_leave","konto","subkonto",
                "benefit_actual","benefit_real","postal_code"
            ])

def log_usage(payload: SimInput, result: dict):
    ensure_log_header()
    with LOG_CSV.open("a", newline="", encoding="utf-8") as f:
        w = csv.writer(f); now = dt.datetime.now()
        w.writerow([
            now.date().isoformat(), now.strftime("%H:%M:%S"),
            payload.expected_pension or "", payload.age, payload.sex.upper(),
            payload.gross_salary, "tak" if payload.include_sick_leave else "nie",
            (payload.zus_balance.konto if payload.zus_balance else 0.0) or 0.0,
            (payload.zus_balance.subkonto if payload.zus_balance else 0.0) or 0.0,
            result["benefit"]["actual"], result["benefit"]["real"],
            payload.postal_code or ""
        ])

def compute_replacement_rate(benefit_real: float, current_gross: float) -> Optional[float]:
    if current_gross > 0:
        return round(100.0 * benefit_real / current_gross, 2)
    return None

def fmt_money(v: Optional[float]) -> str:
    if v is None: return "—"
    return f"{v:,.2f} zł".replace(",", " ").replace("\xa0", " ")

def fmt_pct(v: Optional[float]) -> str:
    if v is None: return "—"
    return f"{v:.2f}%"

def avg_benefit_for_year(year: int, **_ignore) -> Optional[float]:
    if AVG_TABLE:
        if year in AVG_TABLE:
            return AVG_TABLE[year]
        years = sorted(AVG_TABLE.keys())
        y_min, y_max = years[0], years[-1]
        g = _avg_growth_rate_from_tail(AVG_TABLE, k=5)

        if year > y_max:
            v0 = AVG_TABLE[y_max]
            return float(v0 * ((1.0 + g) ** (year - y_max)))
        if year < y_min:
            v0 = AVG_TABLE[y_min]
            return float(v0 / ((1.0 + g) ** (y_min - year)))

    val = ASSUMPTIONS.get("srednia_emerytura_roczna", {}).get(str(year))
    return float(val) if val is not None else None

# --- Bucket specs (pulpit podstawowy) ---
BUCKET_SPECS = [
    {
        "key": "ponizej_min",
        "label": "Poniżej minimalnej",
        "mult": 0.60,
        "tooltip": "Świadczeniobiorcy z krótkim stażem (poniżej 20 lat K / 25 lat M), brak gwarancji minimalnej emerytury."
    },
    {
        "key": "okolice_min",
        "label": "Okolice minimalnej",
        "mult": 0.85,
        "tooltip": "Niski staż/niższe zarobki. Emerytura w okolicach poziomu minimalnego."
    },
    {
        "key": "srednia",
        "label": "Średnia wysokość",
        "mult": 1.00,
        "tooltip": "Najliczniejsza grupa: standardowy staż i typowe wynagrodzenia w trakcie kariery."
    },
    {
        "key": "powyzej_sredniej",
        "label": "Powyżej średniej",
        "mult": 1.30,
        "tooltip": "Dłuższy staż i/lub wyższe wynagrodzenia. Wyższa podstawa do naliczenia świadczenia."
    },
    {
        "key": "najwyzsze",
        "label": "Najwyższe świadczenia",
        "mult": 1.80,
        "tooltip": "Wysoka i stabilna ścieżka wynagrodzeń, długi staż, brak przerw w karierze."
    },
]

def buckets_for_year(year: int):
    """Zwraca buckety oparte o średnią dla danego roku (multiplikatory z BUCKET_SPECS)."""
    avg = avg_benefit_for_year(year) or avg_benefit_for_year(dt.date.today().year) or 4000.0
    out = []
    for b in BUCKET_SPECS:
        out.append({
            "key": b["key"],
            "label": b["label"],
            "amount": round(avg * b["mult"], 2),
            "tooltip": b["tooltip"]
        })
    return {"year": year, "avg_source": "AVG_TABLE/ASSUMPTIONS/DEMO", "buckets": out}

# --- Extra layout knobs (hackathon tuning) ---
KPI_CARD_H = 64     # wysokość małych kart KPI
KPI_GAP    = 18     # odstęp między kartami KPI
COMPARE_H  = 90     # wysokość boxa "Porównanie ze średnią"
NEG_GAP    = 80     # "negatywny" odstęp NAD boxem porównania (dodatnia wartość = podnosi box w górę)

# --- PDF helpers (stare; część zostaje, część poniżej nowa wersja) ---
def draw_header(c: canvas.Canvas, w, h, title: str):
    c.setFillColor(ZUS_NAVY); c.rect(0, h-72, w, 72, fill=1, stroke=0)
    c.setFillColor(ZUS_GREEN); c.rect(0, h-72, 8, 72, fill=1, stroke=0)
    c.setFillColor(colors.white)
    c.setFont(FONT_BOLD, 22); c.drawString(MARGIN_X, h-42, title)
    c.setFont(FONT_MAIN, 10); c.drawString(MARGIN_X, h-64, f"Data: {dt.date.today().isoformat()}")

def draw_footer(c: canvas.Canvas, w):
    c.setFillColor(ZUS_GRAY); c.rect(0, 0, w, 26, fill=1, stroke=0)
    c.setFillColor(ZUS_NAVY); c.setFont(FONT_MAIN, 9)
    c.drawRightString(w-MARGIN_X, 9, f"Emerytura360 • {dt.date.today().isoformat()}")

def section_title(c: canvas.Canvas, text: str, x: float, y: float):
    c.setFillColor(ZUS_NAVY); c.setFont(FONT_BOLD, 12)
    c.drawString(x, y, text)

def wrap_text(c: canvas.Canvas, text: str, font: str, size: int, max_width: float) -> List[str]:
    c.setFont(font, size)
    words = text.split(" ")
    lines: List[str] = []
    cur = ""
    for w in words:
        test = (cur + " " + w).strip()
        if c.stringWidth(test, font, size) <= max_width:
            cur = test
        else:
            if cur:
                lines.append(cur); cur = w
            else:
                while c.stringWidth(w, font, size) > max_width and len(w) > 1:
                    w = w[:-1]
                lines.append(w); cur = ""
    if cur: lines.append(cur)
    if len(lines) > 2:
        lines = [lines[0], lines[1] + "…"]
    return lines

def kpi_card(c: canvas.Canvas, x, y, w, h, label: str, value: str, accent=ZUS_GREEN):
    corner = 12
    c.setFillColor(colors.white)
    c.setStrokeColor(ZUS_GRAY)
    c.setLineWidth(0.8)
    c.roundRect(x, y, w, h, corner, fill=1, stroke=1)

    LABEL_FS = 10
    VALUE_FS = 16
    TOP_PAD = 16
    SIDE_PAD = 12

    c.setFillColor(ZUS_NAVY)
    label_lines = wrap_text(c, label, FONT_MAIN, LABEL_FS, w - 2*SIDE_PAD)
    c.setFont(FONT_MAIN, LABEL_FS)
    base_y = y + h - TOP_PAD
    line_h = LABEL_FS + 2
    for i, line in enumerate(label_lines):
        c.drawCentredString(x + w/2, base_y - i*line_h, line)

    c.setFillColor(ZUS_BLACK)
    c.setFont(FONT_BOLD, VALUE_FS)
    safe_gap = 12
    value_y = base_y - len(label_lines)*line_h - safe_gap
    value_y = max(value_y, y + 10)
    c.drawCentredString(x + w/2, value_y, value)

def comparison_numbers(c: canvas.Canvas, x, y, w,
                       my_value: Optional[float],
                       avg_value: Optional[float]):
    box_h = 60
    c.setFillColor(colors.white); c.setStrokeColor(ZUS_GRAY)
    c.roundRect(x, y, w, box_h, 10, fill=1, stroke=1)

    # ★ Pozycje – środek boxa
    left_x   = x + w * 0.25
    center_x = x + w * 0.50
    right_x  = x + w * 0.75
    mid_y    = y + box_h / 2.0

    LABEL_FS = 9
    VALUE_FS_LEFT  = 14
    VALUE_FS_RIGHT = 14
    LABEL_OFFSET = 10   # ile nad środkiem
    VALUE_OFFSET = 12   # ile pod środkiem

    # Lewa kolumna (Twoja)
    c.setFillColor(ZUS_NAVY); c.setFont(FONT_MAIN, LABEL_FS)
    c.drawCentredString(left_x,  mid_y + LABEL_OFFSET, "Twoja (nominalna, m-c)")
    c.setFillColor(ZUS_BLACK); c.setFont(FONT_BOLD, VALUE_FS_LEFT)
    c.drawCentredString(left_x,  mid_y - VALUE_OFFSET, fmt_money(my_value))

    # Prawa kolumna (Średnia)
    if not avg_value or avg_value <= 0: avg_value = 0.0
    c.setFillColor(ZUS_NAVY); c.setFont(FONT_MAIN, LABEL_FS)
    c.drawCentredString(right_x, mid_y + LABEL_OFFSET, "Średnia")
    c.setFillColor(ZUS_BLACK); c.setFont(FONT_BOLD, VALUE_FS_RIGHT)
    c.drawCentredString(right_x, mid_y - VALUE_OFFSET, fmt_money(avg_value))

    # Badge ze środkiem w boxie
    diff_pct = None
    if avg_value > 0 and my_value is not None:
        diff_pct = (my_value - avg_value) / avg_value * 100.0
    badge_text = "—"; badge_color = ZUS_ORANGE
    if diff_pct is not None:
        badge_text = f"{diff_pct:+.2f}%"
        if diff_pct >= 5: badge_color = ZUS_GREEN
        elif diff_pct <= -5: badge_color = ZUS_RED

    pill_w, pill_h = 50, 16
    pill_x = center_x - pill_w/2
    pill_y = mid_y - pill_h/2 - 20  # ★ idealne wycentrowanie
    c.setFillColor(badge_color); c.setStrokeColor(colors.white); c.setLineWidth(1.2)
    c.roundRect(pill_x, pill_y, pill_w, pill_h, 7, fill=1, stroke=1)
    c.setFillColor(colors.white); c.setFont(FONT_BOLD, 10)
    c.drawCentredString(center_x, pill_y + pill_h/2 - 3, badge_text)

def fmt_money_pl_short(v: float) -> str:
    # skrót „mln/tys.” + polskie separatory
    if v >= 2_000_000:
        s = f"{v/1_000_000:.1f}".replace(".", ",")
        return f"{s} mln zł"
    if v >= 20_000:
        s = str(int(round(v/1000)))  # tysiące bez dziesiętnych
        return f"{s} tys. zł"
    # format z miejscami dziesiętnymi: spacje tysięcy, przecinek dziesiętnych
    s = f"{v:,.2f}".replace(",", "X").replace(".", ",").replace("X", " ")
    return f"{s} zł"


def draw_simple_bar_chart(c: canvas.Canvas, x, y, w, h, labels, values, colors_fill=None):
    n = len(values)
    if n == 0:
        return
    max_v = max(values) if max(values) > 0 else 1.0

    # ramka
    c.setFillColor(colors.white)
    c.setStrokeColor(ZUS_GRAY)
    c.roundRect(x, y, w, h, 10, fill=1, stroke=1)

    # marginesy wewnętrzne (trochę większy top na etykiety)
    pad_x = 24
    pad_top = 32
    pad_bottom = 34
    chart_x = x + pad_x
    chart_w = w - 2*pad_x
    chart_y = y + pad_bottom
    chart_h = h - pad_top - pad_bottom

    # oś
    c.setStrokeColor(ZUS_GRAY)
    c.setLineWidth(0.6)
    c.line(chart_x, chart_y, chart_x + chart_w, chart_y)

    # rozstaw słupków
    gap = chart_w * 0.12 / max(1, n)
    bar_w = (chart_w - gap * (n + 1)) / n

    for i, v in enumerate(values):
        bh = 0 if max_v <= 0 else (v / max_v) * chart_h
        bx = chart_x + gap + i * (bar_w + gap)
        by = chart_y

        # kolor
        if colors_fill and i < len(colors_fill):
            c.setFillColor(colors_fill[i])
        else:
            c.setFillColor(ZUS_BLUE if i == 0 else ZUS_GREEN)

        c.rect(bx, by, bar_w, bh, fill=1, stroke=0)

        # --- Etykieta wartości (PL, skróty) ---
        label = fmt_money_pl_short(v)

        # Dopuszczamy minimalny rozmiar 6 pt.
        # Dodatkowo pozwalamy napisie „wystawać” maksymalnie do połowy szczeliny (gap),
        # żeby uniknąć cięcia na „…”, ale też nie wpadać na sąsiedni słupek.
        value_fs = 8
        max_label_width = bar_w + gap * 0.5
        while c.stringWidth(label, FONT_BOLD, value_fs) > max_label_width and value_fs > 6:
            value_fs -= 1

        # Jedna, stała wysokość dla wszystkich etykiet nad słupkiem
        val_y = min(by + bh + 12, y + h - 10)

        c.setFillColor(ZUS_BLACK)
        c.setFont(FONT_BOLD, value_fs)
        c.drawCentredString(bx + bar_w/2, val_y, label)

        # --- Oś X (co 1, ale będzie miejsce dzięki większym odstępom i rotacji powyżej) ---
        c.setFillColor(ZUS_NAVY)
        c.setFont(FONT_MAIN, 9)
        c.drawCentredString(bx + bar_w/2, y + 6, labels[i])

def bullet_line(c: canvas.Canvas, x, y, text, color=ZUS_GREEN):
    c.setFillColor(color); c.circle(x, y+3, 2.6, fill=1, stroke=0)
    c.setFillColor(ZUS_NAVY); c.setFont(FONT_MAIN, 10)
    c.drawString(x+10, y, text)

# ======== NOWE HELPERY do layoutu „annual report” (prawa szpalta + donuty) ========

def draw_right_panel_bg(c: canvas.Canvas, w, h):
    """Prawa połowa w kolorze ZUS_NAVY."""
    right_x = int(w * 0.5)
    c.setFillColor(ZUS_NAVY)
    c.rect(right_x, 0, w-right_x, h, fill=1, stroke=0)
    c.setFillColor(ZUS_GREEN)
    c.rect(right_x, h-2, w-right_x, 2, fill=1, stroke=0)

def donut(c: canvas.Canvas, cx, cy, r_outer, r_inner, segments, labels=None,
          legend_x=None, legend_y=None, legend_leading=12):
    """
    segments = [(fraction_0to1, color, optional_label)]
    """
    total = sum(s[0] for s in segments) or 1.0
    angle0 = 90.0  # start od góry
    for frac, color_, *_ in segments:
        sweep = 360.0 * (frac/total)
        c.setFillColor(color_); c.setStrokeColor(color_)
        c.wedge(cx-r_outer, cy-r_outer, cx+r_outer, cy+r_outer, angle0, angle0+sweep, stroke=0, fill=1)
        angle0 += sweep
    c.setFillColor(colors.white)
    c.circle(cx, cy, r_inner, fill=1, stroke=0)

    # legenda
    if legend_x is not None and labels:
        c.setFont(FONT_MAIN, 8)
        for i, ((frac, color_, *_), lab) in enumerate(zip(segments, labels)):
            c.setFillColor(color_)
            c.circle(legend_x, legend_y - i*legend_leading + 3, 3, fill=1, stroke=0)
            c.setFillColor(colors.white)
            c.drawString(legend_x + 10, legend_y - i*legend_leading, lab)

def rr_gauge(c: canvas.Canvas, cx: float, cy: float, r_outer: float, r_inner: float,
             rr_frac: float, title: str = "Stopa zastąpienia"):
    """
    Prosty „gauge” w formie pierścienia dla RR.
    rr_frac: 0.0–1.0 (np. 0.1561 dla 15.61%)
    """
    rr = max(0.0, min(1.0, rr_frac))

    # wypełnienie
    c.setFillColor(ZUS_ORANGE)
    c.wedge(cx - r_outer, cy - r_outer, cx + r_outer, cy + r_outer,
            90, 90 + 360 * rr, stroke=0, fill=1)
    c.setFillColor(ZUS_GRAY)
    c.wedge(cx - r_outer, cy - r_outer, cx + r_outer, cy + r_outer,
            90 + 360 * rr, 450, stroke=0, fill=1)

    # dziura w środku (pierścień)
    c.setFillColor(colors.white)
    c.circle(cx, cy, r_inner, fill=1, stroke=0)

    # tytuł + wartość w środku
    c.setFillColor(colors.white); c.setFont(FONT_BOLD, 18)
    c.drawString(cx - r_outer, cy + r_outer + 22, title)

    c.setFillColor(colors.white); c.setFont(FONT_BOLD, 20)
    c.drawCentredString(cx, cy + 2, f"{rr*100:.2f}%")


def section_heading(c: canvas.Canvas, text: str, x: float, y: float, light=False):
    c.setFont(FONT_BOLD, 18)
    c.setFillColor(colors.white if light else ZUS_NAVY)
    c.drawString(x, y, text)

def stat_bar(c: canvas.Canvas, x, y, w, pct, label, bar_color):
    """Pasek postępu w stylu prostych statystyk."""
    h = 14
    c.setFillColor(ZUS_GRAY); c.rect(x, y, w, h, fill=1, stroke=0)
    c.setFillColor(bar_color); c.rect(x, y, w*max(0.0, min(1.0, pct)), h, fill=1, stroke=0)
    c.setFillColor(ZUS_NAVY); c.setFont(FONT_MAIN, 9)
    c.drawString(x, y-16, label)
    c.setFont(FONT_BOLD, 10); c.drawRightString(x+w, y-12, f"{int(max(0,min(100, pct*100)))}%")

def kpi_circle(c: canvas.Canvas, cx, cy, r, title, value, sub=None):
    c.setFillColor(colors.white); c.circle(cx, cy, r, fill=1, stroke=0)
    c.setFillColor(ZUS_NAVY); c.setFont(FONT_MAIN, 9)
    c.drawCentredString(cx, cy+r+10, title)
    c.setFont(FONT_BOLD, 16); c.drawCentredString(cx, cy+2, value)
    if sub:
        c.setFont(FONT_MAIN, 8); c.setFillColor(ZUS_GRAY); c.drawCentredString(cx, cy-12, sub)

# --- API Endpoints ---
@app.get("/assumptions")
def get_assumptions():
    return {
        "today": dt.date.today().isoformat(),
        "assumptions": {
            "cpi_default": 0.03,
            "life_months_default": 240,
            "absencja_chorobowa": ASSUMPTIONS.get("absencja_chorobowa", {}),
            "opoznienie_dodatkowy_wzrost_proc": ASSUMPTIONS.get("opoznienie_dodatkowy_wzrost_proc", {})
        },
        "data_sources": {
            "mentor_params": bool(PARAMS),
            "avg_benefits_file": bool(AVG_TABLE)
        }
    }

@app.post("/simulate")
def simulate(payload: SimInput):
    today = dt.date.today()
    current_year = today.year
    retire_year = payload.retire_year or (today.year + max(0, statutory_retire_age(payload.sex) - payload.age))

    # Walidacje wejścia (przyjazne 400)
    if payload.start_year >= (payload.retire_year or retire_year):
        raise HTTPException(status_code=400, detail="start_year musi być < retire_year")

    if payload.custom_wage_timeline:
        bad = [y for y, v in payload.custom_wage_timeline.items() if v is None or v <= 0]
        if bad:
            raise HTTPException(status_code=400, detail=f"custom_wage_timeline zawiera niepoprawne wartości dla lat: {bad}")

    # === 1) ŚCIEŻKA PŁAC: najpierw spróbuj z arkusza mentorów (PARAMS.avg_wage), potem fallback ===
    def _closest_year(d: Dict[int, dict], target: int) -> Optional[int]:
        if not d:
            return None
        le = [k for k in d.keys() if k <= target]
        if le:
            return max(le)
        return min(d.keys())

    wages: Dict[int, float] = {}
    used_params_path = False

    if not payload.custom_wage_timeline and PARAMS:
        ref_y = _closest_year(PARAMS, current_year)
        have_all = ref_y and all(PARAMS.get(y, {}).get("avg_wage") for y in range(payload.start_year, retire_year))
        if have_all:
            used_params_path = True
            ref_avg = PARAMS[ref_y]["avg_wage"]
            wages = {y: float(payload.gross_salary) * (PARAMS[y]["avg_wage"] / ref_avg)
                     for y in range(payload.start_year, retire_year)}

    if not wages:
        auto_backcast = os.getenv("AUTO_BACKCAST", "1") == "1"
        wg_fallback = wage_growth_rate()
        if not payload.custom_wage_timeline and auto_backcast and payload.start_year < retire_year:
            for y in range(payload.start_year, retire_year):
                years_diff = max(0, current_year - y)
                wages[y] = float(payload.gross_salary) / ((1.0 + wg_fallback) ** years_diff)
        else:
            wages = payload.custom_wage_timeline or {y: payload.gross_salary for y in range(payload.start_year, retire_year)}

    # 1.1. L4 per rok (opcjonalne)
    def l4_factor_for_year(y: int) -> float:
        if payload.custom_sick_days and y in payload.custom_sick_days:
            return efekt_absencji_factor(payload.custom_sick_days[y])
        return efekt_absencji_factor(absencja_days(payload.sex)) if payload.include_sick_leave else 1.0

    # 1.2. Limit 250% przeciętnego wynagrodzenia (miesięczny cap – możesz zostawić lub wyłączyć)
    def _cap_base_monthly(y: int, base: float) -> float:
        avg = PARAMS.get(y, {}).get("avg_wage")
        if avg:
            return min(base, 2.5 * avg)
        return base

    # 1.3. Limit 30× prognozowanej średniej (roczny cap – NOWE)
    def annual_base_with_30x_cap(y: int, monthly_base_after_cap: float) -> float:
        avg = PARAMS.get(y, {}).get("avg_wage")  # prognozowana średnia miesięczna
        annual_base = 12.0 * float(monthly_base_after_cap)
        if avg:
            annual_cap = 30.0 * float(avg)  # 30-krotność miesięcznej średniej = limit roczny
            return min(annual_base, annual_cap)
        return annual_base

    # === 2) Składki roczne po L4 + limitach (250% m-c + 30× rocznie) ===
    skladki_po_latach: Dict[int, float] = {}
    for rok, wyn in wages.items():
        monthly_base = _cap_base_monthly(rok, wyn)             # cap 2.5× m-c (opcjonalny)
        annual_base = annual_base_with_30x_cap(rok, monthly_base)  # cap 30× rocznie
        skladki_po_latach[rok] = annual_base * 0.1952 * l4_factor_for_year(rok)

    # === 3) Waloryzacje i podstawa ===
    rocznie = waloryzuj_rocznie(skladki_po_latach)
    if isinstance(rocznie, dict):
        rocznie_dict = rocznie
    else:
        rocznie_dict = dict(skladki_po_latach)

    po_kwartale = waloryzuj_kwartalnie_po_31_stycznia(retire_year, payload.quarter_award, rocznie)

    konto = (payload.zus_balance.konto if payload.zus_balance else 0.0) or 0.0
    subkonto = (payload.zus_balance.subkonto if payload.zus_balance else 0.0) or 0.0
    podstawa = po_kwartale + konto + subkonto

    # === 4) Annuitetyzacja i urealnienie (CPI z PARAMS jeśli jest) ===
    months = expected_life_months(payload.sex, retire_year)
    benefit_nominal = annuitetyzuj(podstawa, months)

    if PARAMS.get(today.year, {}).get("cpi_index"):
        cpi = max(0.0, (PARAMS[today.year]["cpi_index"] - 1.0))
    else:
        cpi = float(os.getenv("CPI", "0.03"))

    years_to_retire = max(0, retire_year - today.year)
    benefit_real = urealnij(benefit_nominal, cpi, years_to_retire)

    # === 5) Zindeksowane wynagrodzenie do roku przejścia ===
    if PARAMS.get(current_year, {}).get("avg_wage") and PARAMS.get(retire_year, {}).get("avg_wage"):
        indexed_wage_at_retirement = float(payload.gross_salary) * (
            PARAMS[retire_year]["avg_wage"] / PARAMS[current_year]["avg_wage"]
        )
        wg_used = None  # ścieżka z PARAMS
    else:
        wg_used = wage_growth_rate()
        indexed_wage_at_retirement = float(payload.gross_salary) * ((1.0 + wg_used) ** years_to_retire)

    # efektywny CAGR płac (do raportu)
    if wg_used is None:
        if years_to_retire > 0:
            wg_effective = (PARAMS[retire_year]["avg_wage"] / PARAMS[current_year]["avg_wage"]) ** (1.0 / years_to_retire) - 1.0
        else:
            wg_effective = 0.0
    else:
        wg_effective = wg_used

    replacement_rate_indexed = (
        round(100.0 * float(benefit_nominal) / indexed_wage_at_retirement, 2)
        if indexed_wage_at_retirement > 0 else None
    )

    # === 6) Średnia emerytura i stopa zastąpienia (dzisiejsza) ===
    avg_benefit = avg_benefit_for_year(retire_year)
    replacement = compute_replacement_rate(benefit_real, payload.gross_salary)

    # === 7) Referencja: ile byłoby BEZ L4 (komunikat o wpływie absencji) — z tym samym limitem 30× ===
    def _wages_for_range(end_year: int) -> Dict[int, float]:
        if PARAMS and all(PARAMS.get(y, {}).get("avg_wage") for y in range(payload.start_year, end_year)):
            ref_y = _closest_year(PARAMS, current_year) or current_year
            ref_avg = PARAMS[ref_y]["avg_wage"] if PARAMS.get(ref_y, {}) else PARAMS[payload.start_year]["avg_wage"]
            return {y: float(payload.gross_salary) * (PARAMS[y]["avg_wage"] / ref_avg)
                    for y in range(payload.start_year, end_year)}
        wg_tmp = wage_growth_rate()
        w: Dict[int, float] = {}
        for y in range(payload.start_year, end_year):
            years_diff = max(0, current_year - y)
            w[y] = float(payload.gross_salary) / ((1.0 + wg_tmp) ** years_diff)
        return w

    def real_benefit_with_l4_factor(retire_y: int, l4_fact: float) -> float:
        wages_local = payload.custom_wage_timeline or _wages_for_range(retire_y)
        skladki_local = {}
        for rok, wyn in wages_local.items():
            monthly_base = _cap_base_monthly(rok, wyn)
            annual_base = annual_base_with_30x_cap(rok, monthly_base)
            skladki_local[rok] = annual_base * 0.1952 * l4_fact

        rocznie_local = waloryzuj_rocznie(skladki_local)
        po_kw_local = waloryzuj_kwartalnie_po_31_stycznia(retire_y, payload.quarter_award, rocznie_local)
        konto_local = (payload.zus_balance.konto if payload.zus_balance else 0.0) or 0.0
        subkonto_local = (payload.zus_balance.subkonto if payload.zus_balance else 0.0) or 0.0
        podstawa_local = po_kw_local + konto_local + subkonto_local
        months_local = expected_life_months(payload.sex, retire_y)
        nominal_local = annuitetyzuj(podstawa_local, months_local)
        return urealnij(nominal_local, cpi, max(0, retire_y - today.year))

    real_with_L4 = float(benefit_real)
    real_no_L4   = float(real_benefit_with_l4_factor(retire_year, 1.0))
    delta_abs = round(real_no_L4 - real_with_L4, 2)
    delta_pct = round(100.0 * delta_abs / real_no_L4, 2) if real_no_L4 else None

    sick_leave_impact = {
        "real_with_L4": round(real_with_L4, 2),
        "real_without_L4": round(real_no_L4, 2),
        "loss_abs": delta_abs,
        "loss_pct": delta_pct
    }

    # === 8) Goal-seek vs expected_pension ===
    goal_seek = {
        "enabled": False,
        "expected": payload.expected_pension,
        "extra_years_needed": None,
        "target_gap": None,
        "checked_until_year": retire_year
    }
    if payload.expected_pension and payload.expected_pension > 0:
        goal_seek["enabled"] = True
        goal_seek["target_gap"] = round(float(payload.expected_pension - benefit_real), 2)

        MAX_EXTRA = 10
        found = None
        test_year = retire_year
        for add in range(0, MAX_EXTRA + 1):
            y = retire_year + add
            real_y = real_benefit_with_l4_factor(y, l4_factor_for_year(y))
            if real_y >= payload.expected_pension:
                found = add
                test_year = y
                break
            test_year = y

        goal_seek["extra_years_needed"] = found
        goal_seek["checked_until_year"] = test_year

    years_span = list(range(payload.start_year, retire_year))
    if years_span:
        report_l4_factor = sum(l4_factor_for_year(y) for y in years_span) / len(years_span)
    else:
        report_l4_factor = l4_factor_for_year(retire_year)

    # === 9) Wynik ===
    result = {
        "benefit": {"actual": round(float(benefit_nominal), 2), "real": round(float(benefit_real), 2)},
        "retire_year": retire_year,
        "avg_benefit_year": avg_benefit,
        "replacement_rate_percent": replacement,
        "indexed_wage_at_retirement": round(float(indexed_wage_at_retirement), 2),
        "replacement_rate_indexed_percent": replacement_rate_indexed,
        "effect_sick_leave": {"factor": round(float(report_l4_factor), 4)},
        "sick_leave_impact": sick_leave_impact,
        "scenarios": ASSUMPTIONS.get("opoznienie_dodatkowy_wzrost_proc", {}),
        "assumptions_used": {
            "cpi": cpi,
            "life_months": months,
            "wage_growth": wg_effective,
            "wage_growth_source": "mentor_avg_wage" if wg_used is None else "env_fallback"
        },
        "goal_seek": goal_seek,
        "data_sources": {"mentor_params": used_params_path, "avg_benefits_file": bool(AVG_TABLE)}
    }

    # --- log ---
    log_usage(payload, result)
    return result

@app.post("/simulate/timeline")
def simulate_timeline(payload: SimInput, format: Optional[str] = Query(None)):
    """
    Zwraca roczny timeline dla dashboardu:
    - year
    - base_after_indexation (po waloryzacji kwartalnej w danym roku)
    - benefit_if_retire_in_year: {nominal, real}
    """
    _fmt = format if isinstance(format, (str, type(None))) else None
    today = dt.date.today()
    current_year = today.year
    start_y = payload.start_year
    default_retire_year = today.year + max(0, statutory_retire_age(payload.sex) - payload.age)
    end_y = payload.retire_year or default_retire_year
    retire_year = payload.retire_year or default_retire_year

    # Walidacje wejścia (przyjazne 400)
    if payload.start_year >= (payload.retire_year or retire_year):
        raise HTTPException(status_code=400, detail="start_year musi być < retire_year")

    if payload.custom_wage_timeline:
        bad = [y for y, v in payload.custom_wage_timeline.items() if v is None or v <= 0]
        if bad:
            raise HTTPException(status_code=400, detail=f"custom_wage_timeline zawiera niepoprawne wartości dla lat: {bad}")

    # 1) Ścieżka płac jak w simulate: PARAMS -> fallback
    def _closest_year(d: Dict[int, dict], target: int) -> Optional[int]:
        if not d:
            return None
        le = [k for k in d.keys() if k <= target]
        if le:
            return max(le)
        return min(d.keys())

    wages: Dict[int, float] = {}
    if not payload.custom_wage_timeline and PARAMS:
        ref_y = _closest_year(PARAMS, current_year)
        have_all = ref_y and all(PARAMS.get(y, {}).get("avg_wage") for y in range(start_y, end_y))
        if have_all:
            ref_avg = PARAMS[ref_y]["avg_wage"]
            wages = {y: float(payload.gross_salary) * (PARAMS[y]["avg_wage"] / ref_avg)
                     for y in range(start_y, end_y)}

    if not wages:
        auto_backcast = os.getenv("AUTO_BACKCAST", "1") == "1"
        wg = wage_growth_rate()
        if not payload.custom_wage_timeline and auto_backcast and start_y < end_y:
            for y in range(start_y, end_y):
                years_diff = max(0, current_year - y)
                wages[y] = float(payload.gross_salary) / ((1.0 + wg) ** years_diff)
        else:
            wages = payload.custom_wage_timeline or {y: payload.gross_salary for y in range(start_y, end_y)}

    # 2) L4 per rok
    def l4_factor_for_year(y: int) -> float:
        if payload.custom_sick_days and y in payload.custom_sick_days:
            return efekt_absencji_factor(payload.custom_sick_days[y])
        return efekt_absencji_factor(absencja_days(payload.sex)) if payload.include_sick_leave else 1.0

    # 3) Limit 250% przeciętnego wynagrodzenia (miesięczny cap – opcjonalny)
    def _cap_base_monthly(y: int, base: float) -> float:
        avg = PARAMS.get(y, {}).get("avg_wage")
        if avg:
            return min(base, 2.5 * avg)
        return base

    # 3b) Limit 30× prognozowanej średniej (roczny cap – NOWE)
    def annual_base_with_30x_cap(y: int, monthly_base_after_cap: float) -> float:
        avg = PARAMS.get(y, {}).get("avg_wage")
        annual_base = 12.0 * float(monthly_base_after_cap)
        if avg:
            annual_cap = 30.0 * float(avg)
            return min(annual_base, annual_cap)
        return annual_base

    # 4) CPI dla urealnienia (z PARAMS jeżeli jest)
    if PARAMS.get(today.year, {}).get("cpi_index"):
        cpi = max(0.0, (PARAMS[today.year]["cpi_index"] - 1.0))
    else:
        cpi = float(os.getenv("CPI", "0.03"))

    out: List[Dict] = []
    base_running = 0.0  # skumulowana baza po WALORYZACJI ROCZNEJ do końca danego roku

    for y in range(start_y, end_y):
        wage_y = wages.get(y, payload.gross_salary)
        monthly_base = _cap_base_monthly(y, wage_y)
        annual_base = annual_base_with_30x_cap(y, monthly_base)
        contr_y = annual_base * 0.1952 * l4_factor_for_year(y)

        rocznie = waloryzuj_rocznie({y: contr_y})
        if isinstance(rocznie, dict):
            add_indexed = float(sum(rocznie.values()))
        else:
            add_indexed = float(rocznie)

        base_running += add_indexed

        base_after_q = waloryzuj_kwartalnie_po_31_stycznia(y, payload.quarter_award, base_running)

        konto = (payload.zus_balance.konto if payload.zus_balance else 0.0) or 0.0
        subkonto = (payload.zus_balance.subkonto if payload.zus_balance else 0.0) or 0.0
        podstawa_y = base_after_q + konto + subkonto

        months = expected_life_months(payload.sex, y)
        nominal = annuitetyzuj(podstawa_y, months)
        real = urealnij(nominal, cpi, max(0, y - dt.date.today().year))

        out.append({
            "year": y,
            "base_after_indexation": round(float(podstawa_y), 2),
            "benefit_if_retire_in_year": {
                "nominal": round(float(nominal), 2),
                "real": round(float(real), 2),
            }
        })

    if (_fmt or "").lower() == "csv":
        import csv as _csv, io as _io
        buf = _io.StringIO()
        w = _csv.writer(buf)
        w.writerow(["year", "base_after_indexation", "benefit_nominal", "benefit_real"])
        for row in out:
            w.writerow([
                row["year"],
                row["base_after_indexation"],
                row["benefit_if_retire_in_year"]["nominal"],
                row["benefit_if_retire_in_year"]["real"]
            ])
        return Response(content=buf.getvalue(), media_type="text/csv")
    return {"timeline": out}

@app.post("/simulate/what-if")
def simulate_what_if(
    payload: SimInput,
    delays: List[int] = Query([0, 1, 2, 5], description="Lata opóźnienia vs retire_year")
):
    """
    Zwraca zestaw scenariuszy dla różnych opóźnień przejścia (0,1,2,5 lat).
    Bazuje na /simulate – więc używa arkusza mentorów, limitu 250% itd.
    """
    # baseline
    base = simulate(payload)
    out = []
    for d in delays:
        p2 = payload.model_copy(update={"retire_year": (payload.retire_year or base["retire_year"]) + d})
        sim_d = simulate(p2)
        out.append({
            "delay_years": d,
            "retire_year": sim_d["retire_year"],
            "benefit": sim_d["benefit"],
            "replacement_rate_percent": sim_d["replacement_rate_percent"],
            "replacement_rate_indexed_percent": sim_d["replacement_rate_indexed_percent"],
            "sick_leave_impact": sim_d["sick_leave_impact"],
            "avg_benefit_year": sim_d["avg_benefit_year"],
            "assumptions_used": sim_d["assumptions_used"]
        })

    # różnice vs baseline (dla wygody frontu)
    for row in out:
        row["delta_vs_baseline"] = {
            "benefit_actual": round(row["benefit"]["actual"] - base["benefit"]["actual"], 2),
            "benefit_real": round(row["benefit"]["real"] - base["benefit"]["real"], 2),
            "replacement_rate_pp": (
                None if base["replacement_rate_percent"] is None or row["replacement_rate_percent"] is None
                else round(row["replacement_rate_percent"] - base["replacement_rate_percent"], 2)
            )
        }

    return {
        "baseline_retire_year": base["retire_year"],
        "baseline_benefit": base["benefit"],
        "scenarios": out
    }

@app.post("/simulate/explain")
def simulate_explain(payload: SimInput):
    """
    Zwraca breakdown: składki roczne (po L4 i limicie), suma po waloryzacji rocznej,
    baza po waloryzacji kwartalnej, annuitetyzację i urealnienie.
    """
    # powtórzenie kluczowych fragmentów z /simulate (lokalnie, bez loga do CSV)
    today = dt.date.today()
    current_year = today.year
    retire_year = payload.retire_year or (today.year + max(0, statutory_retire_age(payload.sex) - payload.age))

    def _closest_year(d: Dict[int, dict], target: int) -> Optional[int]:
        if not d: return None
        le = [k for k in d.keys() if k <= target]
        return max(le) if le else min(d.keys())

    # ścieżka płac (PARAMS -> fallback)
    wages: Dict[int, float] = {}
    ref_y = _closest_year(PARAMS, current_year) if PARAMS else None
    if not payload.custom_wage_timeline and PARAMS and ref_y and all(PARAMS.get(y,{}).get("avg_wage") for y in range(payload.start_year, retire_year)):
        ref_avg = PARAMS[ref_y]["avg_wage"]
        wages = {y: float(payload.gross_salary) * (PARAMS[y]["avg_wage"] / ref_avg)
                 for y in range(payload.start_year, retire_year)}
    else:
        auto_backcast = os.getenv("AUTO_BACKCAST", "1") == "1"
        wg = wage_growth_rate()
        if not payload.custom_wage_timeline and auto_backcast and payload.start_year < retire_year:
            for y in range(payload.start_year, retire_year):
                years_diff = max(0, current_year - y)
                wages[y] = float(payload.gross_salary) / ((1.0 + wg) ** years_diff)
        else:
            wages = payload.custom_wage_timeline or {y: payload.gross_salary for y in range(payload.start_year, retire_year)}

    def l4_factor_for_year(y: int) -> float:
        if payload.custom_sick_days and y in payload.custom_sick_days:
            return efekt_absencji_factor(payload.custom_sick_days[y])
        return efekt_absencji_factor(absencja_days(payload.sex)) if payload.include_sick_leave else 1.0

    def _cap_base(y: int, base: float) -> float:
        avg = PARAMS.get(y, {}).get("avg_wage")
        return min(base, 2.5 * avg) if avg else base

    # 1) składki per rok (po L4 i limicie)
    skladki_po_latach = {}
    per_year = []
    for rok, wyn in wages.items():
        base_y = _cap_base(rok, wyn)
        l4f = l4_factor_for_year(rok)
        contr = base_y * 0.1952 * l4f
        skladki_po_latach[rok] = contr
        per_year.append({"year": rok, "wage": round(wyn,2), "base_after_cap": round(base_y,2), "l4_factor": round(l4f,4), "contribution": round(contr,2)})

    # 2) waloryzacja roczna (sumarycznie)
    rocznie = waloryzuj_rocznie(skladki_po_latach)
    if isinstance(rocznie, dict):
        base_after_annual = float(sum(rocznie.values()))
    else:
        base_after_annual = float(rocznie)

    # 3) waloryzacja kwartalna do roku przejścia
    base_after_quarter = waloryzuj_kwartalnie_po_31_stycznia(retire_year, payload.quarter_award, base_after_annual)

    # 4) podstawa + annuitetyzacja + urealnienie
    konto = (payload.zus_balance.konto if payload.zus_balance else 0.0) or 0.0
    subkonto = (payload.zus_balance.subkonto if payload.zus_balance else 0.0) or 0.0
    podstawa = base_after_quarter + konto + subkonto

    months = expected_life_months(payload.sex, retire_year)
    nominal = annuitetyzuj(podstawa, months)

    cpi = max(0.0, (PARAMS[today.year]["cpi_index"] - 1.0)) if PARAMS.get(today.year,{}).get("cpi_index") else float(os.getenv("CPI","0.03"))
    years_to_retire = max(0, retire_year - today.year)
    real = urealnij(nominal, cpi, years_to_retire)

    return {
        "retire_year": retire_year,
        "step_by_step": {
            "per_year": per_year,
            "sum_after_annual_indexation": round(base_after_annual, 2),
            "base_after_quarter_indexation": round(float(base_after_quarter), 2),
            "konto": round(konto,2),
            "subkonto": round(subkonto,2),
            "podstawa": round(float(podstawa),2),
            "months": months,
            "benefit_nominal": round(float(nominal),2),
            "cpi_used": cpi,
            "benefit_real": round(float(real),2)
        }
    }

@app.get("/buckets")
def get_buckets(year: Optional[int] = None):
    """
    Buckety do pulpitu podstawowego.
    Jeśli `year` nie podany -> bierze bieżący rok.
    """
    y = year or dt.date.today().year
    return buckets_for_year(y)

# ---------- PODMIANA LAYOUTU PDF pod TYM SAMYM endpointem /report/pdf ----------
@app.post(
    "/report/pdf",
    responses={200: {"content": {"application/pdf": {"schema": {"type":"string","format":"binary"}}},
                     "description":"Pobierz wygenerowany raport PDF"}}
)
def report_pdf(payload: SimInput = Body(...)):
    result = simulate(payload)

    buffer = io.BytesIO()
    pdf_name = "raport_emerytura.pdf"
    c = canvas.Canvas(buffer, pagesize=A4)
    c.setTitle(pdf_name)
    w, h = A4

    # --- PANEL: pełna szerokość (z marginesami) ---
    panel_x = MARGIN_X
    panel_w = w - 2 * MARGIN_X

    # Tytuł
    title_y = h - 96
    c.setFillColor(ZUS_NAVY)
    c.setFont(FONT_BOLD, 30); c.drawString(panel_x, title_y, "Raport 2025")
    c.setFont(FONT_BOLD, 22); c.drawString(panel_x, title_y - 28, "Prognoza Emerytury")

    # KPI (2 kolumny nad całą stroną)
    y = title_y - 70
    kpi_gap = 64
    kpi_left_x  = panel_x
    kpi_right_x = panel_x + panel_w/2 + 24

    c.setFont(FONT_MAIN, 10); c.setFillColor(ZUS_NAVY)
    c.drawString(kpi_left_x,  y, "Nominalne (m-c)")
    c.drawString(kpi_right_x, y, "Realne dziś (m-c)")

    y -= 20
    c.setFont(FONT_BOLD, 18); c.setFillColor(ZUS_BLACK)
    c.drawString(kpi_left_x,  y, fmt_money(result['benefit']['actual']))
    c.drawString(kpi_right_x, y, fmt_money(result['benefit']['real']))

    # --- Dwie kolumny treści głównej ---
    GUTTER = 28
    col_w = (panel_w - GUTTER) / 2
    left_x  = panel_x
    right_x = panel_x + col_w + GUTTER

    # Stałe odstępy
    HEADING_GAP   = 36
    SECTION_GAP_Y = 16
    BAR_SPACING   = 34

    # Start kolumn pod KPI
    left_y  = y - HEADING_GAP
    right_y = y - HEADING_GAP

    # ===== LEWA KOLUMNA: „Najważniejsze wskaźniki” (4 paski) =====
    section_heading(c, "Najważniejsze wskaźniki", left_x, left_y)
    left_y -= SECTION_GAP_Y + 16

    # 1) Utrata z L4
    imp = result.get("sick_leave_impact", {})
    loss_pct = max(0.0, min(1.0, (imp.get("loss_pct") or 0) / 100.0))
    stat_bar(c, left_x, left_y, col_w, loss_pct, "Utrata z L4 (proc.)", ZUS_ORANGE)
    left_y -= BAR_SPACING

    # 2) Progres vs oczekiwania
    gs = result.get("goal_seek", {})
    tgt = float(gs.get("expected") or 0)
    prog = float(result['benefit']['real'] or 0)
    pct_goal = 0 if tgt <= 0 else max(0.0, min(1.0, prog / tgt))
    stat_bar(c, left_x, left_y, col_w, pct_goal, "Progres względem oczekiwań", ZUS_GREEN)
    left_y -= BAR_SPACING

    # 3) Twoja nominalna vs średnia
    avg_nom = float(result.get("avg_benefit_year") or 0)
    pct_vs_avg = 0 if avg_nom <= 0 else max(0.0, min(1.0, float(result['benefit']['actual']) / avg_nom))
    stat_bar(c, left_x, left_y, col_w, pct_vs_avg, "Twoja nominalna vs średnia", ZUS_BLUE)
    left_y -= BAR_SPACING

    # 4) RR – stopa zastąpienia
    rr_pct = max(0.0, min(1.0, (result.get("replacement_rate_percent") or 0) / 100.0))
    stat_bar(c, left_x, left_y, col_w, rr_pct, "Stopa zastąpienia (RR)", ACCENT_TEAL)
    left_y -= (BAR_SPACING + 6)

    # ===== PRAWA KOLUMNA: Porównanie + Założenia =====
    section_heading(c, "Porównanie ze średnią", right_x, right_y)
    right_y -= (SECTION_GAP_Y + 4)

    # Box z liczbami (stała wysokość 65)
    cmp_h = 65
    comparison_numbers(
        c,
        x=right_x,
        y=right_y - cmp_h,  # y w tej funkcji to dolna krawędź
        w=col_w,
        my_value=float(result["benefit"]["actual"]),
        avg_value=avg_nom,
    )
    right_y -= (cmp_h + 24)

    # ===== DANE WEJŚCIOWE (lewa) + ZAŁOŻENIA (prawa) – DWIE KOLUMNY =====
    # Punkt startu sekcji pod wcześniejszymi kolumnami
    y_after_cols = min(left_y, right_y) - 18

    # Szerokości/pozycje kolumn w tym rzędzie
    col2_w   = (panel_w - GUTTER) / 2
    left2_x  = panel_x
    right2_x = panel_x + col2_w + GUTTER

    # --- Nagłówki obu kolumn
    section_heading(c, "Dane wejściowe", left2_x, y_after_cols)
    section_heading(c, "Założenia i parametry", right2_x, y_after_cols)

    # Ustawienia typografii list
    LINE_H = 12
    c.setFont(FONT_MAIN, 9)
    c.setFillColor(ZUS_NAVY)

    # --- LEWA kolumna: Dane wejściowe
    left_list_y = y_after_cols - 16
    bullets = [
        f"Wiek: {payload.age}",
        f"Płeć: {payload.sex.upper()}",
        f"Pensja brutto (dziś): {fmt_money(payload.gross_salary)}",
        f"Lata pracy: {payload.start_year}–{result['retire_year']}",
        f"Kwartał przyznania: {payload.quarter_award}",
        f"Uwzględniono L4: {'tak' if payload.include_sick_leave else 'nie'}",
        f"Konto/Subkonto: {fmt_money((payload.zus_balance.konto if payload.zus_balance else 0.0))} / "
            f"{fmt_money((payload.zus_balance.subkonto if payload.zus_balance else 0.0))}",
    ]
    if payload.expected_pension:
        expected = float(payload.expected_pension or 0)
        prognoza_real = float(result["benefit"]["real"] or 0)  # prognoza: realne dziś (m-c)
        gap = max(0.0, round(expected - prognoza_real, 2))

        gs = result.get("goal_seek", {}) or {}
        # jeśli nie osiągamy celu w limicie (MAX_EXTRA=10), pokaż ">10"
        if gs.get("enabled"):
            yn = gs.get("extra_years_needed")
            years_label = ">10" if yn is None else str(int(yn))
        else:
            years_label = "—"

        bullets.extend([
            f"Oczekiwana emerytura: {fmt_money(expected)}",
            f"Oczekiwana vs prognoza: {fmt_money(expected)} vs {fmt_money(prognoza_real)}",
            f"Brakuje do oczekiwań: {fmt_money(gap)}",
            f"Lata potrzebne: {years_label}",
        ])

    for line in bullets:
        c.drawString(left2_x, left_list_y, f"• {line}")
        left_list_y -= LINE_H

    # --- PRAWA kolumna: Założenia i parametry
    right_list_y = y_after_cols - 16
    wg_src = "Mentor (średnia płaca)" if result['assumptions_used']['wage_growth_source']=='mentor_avg_wage' \
            else "ENV fallback"
    assumptions_lines = [
        f"CPI użyte: {fmt_pct((result['assumptions_used']['cpi'] or 0)*100)}",
        f"Oczek. długość życia: {result['assumptions_used']['life_months']} mies.",
        f"CAGR płac do przejścia: {fmt_pct((result['assumptions_used']['wage_growth'] or 0)*100)} ({wg_src})",
        f"Rok przejścia: {result['retire_year']}",
        f"Średnia emerytura w roku przejścia: {fmt_money(float(result.get('avg_benefit_year') or 0))}",
        f"RR (dziś): {fmt_pct((result.get('replacement_rate_percent') or 0))}",
    ]
    # ▼ DODAJ: jawna informacja o średniej absencji (dni) i wpływie na świadczenie
    dni_l4 = ASSUMPTIONS.get("absencja_chorobowa", {}).get(payload.sex.upper(), {}).get("dni_rocznie")
    loss_abs = (result.get("sick_leave_impact", {}) or {}).get("loss_abs")
    loss_pct = (result.get("sick_leave_impact", {}) or {}).get("loss_pct")
    if dni_l4 is not None and loss_pct is not None:
        assumptions_lines.append(
            f"Absencja: ~{int(dni_l4)} dni/rok; wpływ: {fmt_money(loss_abs)} ({fmt_pct(loss_pct)})"
        )

    for line in assumptions_lines:
        c.drawString(right2_x, right_list_y, f"• {line}")
        right_list_y -= LINE_H


    # Stopka
    c.setFillColor(ZUS_GRAY); c.rect(0, 0, w, 26, fill=1, stroke=0)
    c.setFillColor(ZUS_NAVY); c.setFont(FONT_MAIN, 9)
    c.drawRightString(w - MARGIN_X, 9, f"Emerytura360 • {dt.date.today().isoformat()}")

    # --- STRONA 2: Wykresy z /simulate/timeline ---
    # Zakończ stronę 1 i zacznij nową
    c.showPage()

    # Nagłówki strony 2
    w, h = A4
    panel_x = MARGIN_X
    panel_w = w - 2 * MARGIN_X
    c.setFillColor(ZUS_NAVY)
    c.setFont(FONT_BOLD, 24); c.drawString(panel_x, h - 96, "Jak rośnie Twoja emerytura")
    c.setFont(FONT_BOLD, 14); c.drawString(panel_x, h - 120, "Podstawa (konto+subkonto+waloryzacje) i świadczenie realne")

    # 1) Pobierz timeline (lokalnie, bez HTTP) i wybierz kilka równych punktów na osi czasu
    tl = simulate_timeline(payload, format=None).get("timeline", [])
    if tl:
        target_points = 8
        step = max(1, len(tl) // target_points)
        sample = [tl[i] for i in range(0, len(tl), step)]
        if sample[-1]["year"] != tl[-1]["year"]:
            sample.append(tl[-1])

        labels = [str(r["year"]) for r in sample]
        values_base = [float(r["base_after_indexation"]) for r in sample]
        values_real = [float(r["benefit_if_retire_in_year"]["real"]) for r in sample]

        chart_w = panel_w
        chart_h = 160
        x = panel_x
        y_top = h - 170

        # a) Baza po waloryzacjach
        c.setFont(FONT_BOLD, 12); c.setFillColor(ZUS_NAVY)
        c.drawString(x, y_top, "Podstawa po waloryzacjach (koniec roku)")
        draw_simple_bar_chart(
            c,
            x=x,
            y=y_top - (chart_h + 8),
            w=chart_w,
            h=chart_h,
            labels=labels,
            values=values_base,
            colors_fill=[ZUS_BLUE] * len(values_base)
        )

        # b) Świadczenie realne (m-c)
        y2 = y_top - (chart_h + 8) - 40 - chart_h
        c.setFont(FONT_BOLD, 12); c.setFillColor(ZUS_NAVY)
        c.drawString(x, y2 + chart_h + 8, "Prognozowane świadczenie REAL (m-c) – punktowo po latach")
        draw_simple_bar_chart(
            c,
            x=x,
            y=y2,
            w=chart_w,
            h=chart_h,
            labels=labels,
            values=values_real,
            colors_fill=[ZUS_GREEN] * len(values_real)
        )

        c.setFont(FONT_MAIN, 9); c.setFillColor(ZUS_NAVY)
        c.drawString(x, y2 - 18, "Kwoty w PLN (miesięcznie)")
    else:
        c.setFont(FONT_BOLD, 14); c.setFillColor(ZUS_RED)
        c.drawString(panel_x, h - 160, "Brak danych do wykresu timeline.")

    # Stopka strony 2
    c.setFillColor(ZUS_GRAY); c.rect(0, 0, w, 26, fill=1, stroke=0)
    c.setFillColor(ZUS_NAVY); c.setFont(FONT_MAIN, 9)
    c.drawRightString(w - MARGIN_X, 9, f"Emerytura360 • {dt.date.today().isoformat()}")


    c.save(); buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={pdf_name}"}
    )

@app.get(
    "/report/pdf/example",
    responses={200: {"content": {"application/pdf": {"schema": {"type":"string","format":"binary"}}},
                     "description":"Przykładowy raport PDF bez podawania payloadu"}}
)
def report_pdf_example():
    sample = SimInput(
        age=28, sex="K", gross_salary=8500,
        start_year=2020, retire_year=2065,
        include_sick_leave=True, quarter_award=3,
        zus_balance=Balance(konto=0, subkonto=0),
        custom_wage_timeline=None,
        expected_pension=5000, postal_code="30-001"
    )
    return report_pdf(sample)

@app.get(
    "/admin/export-xls",
    responses={200: {"content": {"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
                                 {"schema": {"type":"string","format":"binary"}}},
                     "description":"Eksport użyć symulatora (XLSX) – z logów"}}
)
def export_xls():
    ensure_log_header()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Użycia"
    headers = ["Data użycia","Godzina użycia","Emerytura oczekiwana","Wiek","Płeć","Wynagrodzenie",
               "Czy uwzględniał okresy choroby","Środki konto","Środki subkonto",
               "Emerytura rzeczywista","Emerytura urealniona","Kod pocztowy"]
    ws.append(headers)
    with LOG_CSV.open("r", newline="", encoding="utf-8") as f:
        r = csv.DictReader(f)
        for row in r:
            ws.append([row["date"],row["time"],row["expected_pension"],row["age"],row["sex"],row["salary"],
                       row["included_sick_leave"],row["konto"],row["subkonto"],
                       row["benefit_actual"],row["benefit_real"],row["postal_code"]])
    tmp = io.BytesIO(); wb.save(tmp); tmp.seek(0)
    return StreamingResponse(tmp,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":"attachment; filename=uzycia_symulatora.xlsx"})

@app.post("/admin/clear-logs")
def clear_logs():
    # usuń plik i zainicjalizuj nagłówek ponownie
    try:
        LOG_CSV.unlink(missing_ok=True)  # Python 3.8+: missing_ok
    except TypeError:
        # fallback dla starszych pythonów
        if LOG_CSV.exists():
            LOG_CSV.unlink()
    ensure_log_header()
    return {"cleared": True}

@app.post("/admin/reload")
def reload_tables():
    PARAMS.clear()
    AVG_TABLE.clear()
    load_params_table()
    load_avg_benefit_table()
    return {
        "reloaded": True,
        "params_loaded": bool(PARAMS),
        "avg_loaded": bool(AVG_TABLE),
        "params_years": [min(PARAMS.keys()), max(PARAMS.keys())] if PARAMS else [],
        "avg_years": [min(AVG_TABLE.keys()), max(AVG_TABLE.keys())] if AVG_TABLE else []
    }

@app.get("/admin/sources")
def sources():
    # skrócony podgląd co mamy w pamięci
    sample_params = None
    if PARAMS:
        y = sorted(PARAMS.keys())[0]
        sample_params = {k: v for k, v in PARAMS[y].items() if v is not None}

    return {
        "params_loaded": bool(PARAMS),
        "avg_loaded": bool(AVG_TABLE),
        "params_years_range": [min(PARAMS.keys()), max(PARAMS.keys())] if PARAMS else [],
        "avg_years_range": [min(AVG_TABLE.keys()), max(AVG_TABLE.keys())] if AVG_TABLE else [],
        "sample_params_first_year": sample_params
    }
